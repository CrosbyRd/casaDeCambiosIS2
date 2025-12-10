#reportes/views.py
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.auth.decorators import login_required
from monedas.models import Moneda
from transacciones.models import Transaccion
from reportlab.lib.pagesizes import letter, landscape
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render
from decimal import Decimal
from django.template.defaultfilters import floatformat
from django.utils.formats import number_format


# =========================
# PANEL PRINCIPAL DE REPORTES
# =========================
@login_required
def panel_reportes(request):
    return render(request, 'reportes/panel_reportes.html')


# =========================
# REPORTE DE GANANCIAS (WEB)
# =======================
from django.contrib.auth.decorators import login_required, user_passes_test
from transacciones.models import Transaccion
from monedas.models import Moneda
from datetime import datetime
from decimal import Decimal

from ganancias.models import RegistroGanancia

@login_required
def reporte_ganancias(request):
    # 📌 Trae TODAS las transacciones COMPLETADAS (compra y venta)
    transacciones = Transaccion.objects.filter(estado='completada').order_by('-fecha_creacion')

    # --- FILTROS  ---
    tipo = request.GET.get('tipo')
    if tipo in ['compra', 'venta']:
        transacciones = transacciones.filter(tipo_operacion=tipo)

    moneda = request.GET.get('moneda')
    if moneda and moneda != "todas":
        transacciones = transacciones.filter(
            Q(moneda_origen__codigo=moneda) | Q(moneda_destino__codigo=moneda)
        )

    fecha_desde = request.GET.get('fecha_inicio')
    fecha_hasta = request.GET.get('fecha_fin')

    if fecha_desde:
        transacciones = transacciones.filter(fecha_creacion__date__gte=fecha_desde)

    if fecha_hasta:
        transacciones = transacciones.filter(fecha_creacion__date__lte=fecha_hasta)
    cliente = request.GET.get('cliente')
    if cliente:
        transacciones = transacciones.filter(cliente__nombre__icontains=cliente)
    # --- CÁLCULOS ---
    total_ventas = Decimal('0')
    total_compras = Decimal('0')
    total_general = Decimal('0')
    total_transacciones = transacciones.count()

   # dentro del loop de cálculo de ganancias
    for t in transacciones:
        try:
            registro = RegistroGanancia.objects.get(transaccion=t)
            ganancia = registro.ganancia_registrada
        except RegistroGanancia.DoesNotExist:
            ganancia = Decimal('0')

        t.ganancia = ganancia  # se muestra luego en tabla
        t.ganancia_negativa = ganancia < 0  # <-- nueva variable para template

        if t.tipo_operacion == "venta":
            total_ventas += ganancia
        elif t.tipo_operacion == "compra":
            total_compras += ganancia

        total_general += ganancia


    # -----------------------------
    # 📌 PAGINACIÓN 
    # -----------------------------
    paginator = Paginator(transacciones, 10)  # ← 20 items por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 
    # -----------------------------

    context = {
        'user': request.user,
        'now': datetime.now(),
        'transacciones': page_obj,   # ← AHORA ESTA PAGINADO
        'page_obj': page_obj,        # ← NECESARIO PARA los botones
        'total_ventas': total_ventas,
        'total_compras': total_compras,
        'total_ganancia': total_general,
        'total_transacciones': total_transacciones,
        'monedas': Moneda.objects.all().order_by('codigo'),
    }

    return render(request, 'reportes/reporte_ganancias.html', context)

# =========================
# REPORTE DE GANANCIAS PDF 
# =========================
@login_required
def reporte_ganancias_pdf(request):

    transacciones = Transaccion.objects.filter(estado='completada').order_by('-fecha_creacion')

    # --- FILTROS ---
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')
    cliente = request.GET.get('cliente')
    moneda = request.GET.get('moneda')

    if fecha_inicio:
        transacciones = transacciones.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        transacciones = transacciones.filter(fecha_creacion__date__lte=fecha_fin)
    if tipo:
        transacciones = transacciones.filter(tipo_operacion=tipo)
    if cliente:
        transacciones = transacciones.filter(cliente__nombre__icontains=cliente)
    if moneda:
        transacciones = transacciones.filter(
            Q(moneda_origen__codigo=moneda) |
            Q(moneda_destino__codigo=moneda)
        )

    # --- PDF ---
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ganancias.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),
        rightMargin=20, leftMargin=20,
        topMargin=60, bottomMargin=40
    )

    styles = getSampleStyleSheet()
    elements = []

    user = request.user
    nombre_usuario = getattr(user, "nombre", None) or user.email

    elements.append(Paragraph("<b>Reporte de Ganancias - Global Exchange</b>", styles["Title"]))
    elements.append(Paragraph(f"Generado por: {nombre_usuario}", styles["Normal"]))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elements.append(Spacer(1, 18))

    # Encabezado tabla
    data = [["#", "Cliente", "Tipo", "Moneda","Tasa aplicada", "Monto","Comisión", "Ganancia (Gs)", "Fecha"]]

    total_ventas = Decimal('0')
    total_compras = Decimal('0')
    total_general = Decimal('0')

    # ============================
    # RECORRER TRANSACCIONES
    # ============================
    for idx, t in enumerate(transacciones, start=1):

        # Ganancia registrada
        try:
            registro = RegistroGanancia.objects.get(transaccion=t)
            ganancia = registro.ganancia_registrada
        except RegistroGanancia.DoesNotExist:
            ganancia = Decimal('0')

        moneda_codigo = (
            t.moneda_destino.codigo if t.tipo_operacion == "venta"
            else t.moneda_origen.codigo
        )

        monto = (
            t.monto_destino if t.tipo_operacion == "venta"
            else t.monto_origen
        )

        # Acumular totales
        if t.tipo_operacion == "venta":
            total_ventas += ganancia
        else:
            total_compras += ganancia

        total_general += ganancia

        # Agregar fila
        data.append([
            idx,
            str(t.cliente),
            t.get_tipo_operacion_display(),
            moneda_codigo,
            f"{Decimal(t.tasa_cambio_aplicada):,.0f}".replace(',', '.'),
            f"{Decimal(monto):,.0f}".replace(',', '.'),
            f"{Decimal(t.comision_final):,.0f}".replace(',', '.'),
            f"{ganancia:,.0f}".replace(',', '.'),
            t.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])

    # ============================
    # TOTALES
    # ============================
    data.append(["", "", "", "", "","", "Total Ventas:", f"{total_ventas:,.0f}".replace(',', '.'), ""])
    data.append(["", "", "", "", "","", "Total Compras:", f"{total_compras:,.0f}".replace(',', '.'), ""])
    data.append(["", "", "", "", "", "","Total General:", f"{total_general:,.0f}".replace(',', '.'), ""])


    table = Table(data, hAlign="CENTER")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elements.append(table)
    doc.build(elements)
    return response

# =========================
# REPORTE DE GANANCIAS EXCEL 
# =========================
from openpyxl.utils import get_column_letter

@login_required
def reporte_ganancias_excel(request):

    transacciones = Transaccion.objects.filter(estado='completada').order_by('-fecha_creacion')

    # --- FILTROS ---
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')
    cliente = request.GET.get('cliente')
    moneda = request.GET.get('moneda')

    if fecha_inicio:
        transacciones = transacciones.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        transacciones = transacciones.filter(fecha_creacion__date__lte=fecha_fin)
    if tipo:
        transacciones = transacciones.filter(tipo_operacion=tipo)
    if cliente:
        transacciones = transacciones.filter(cliente__nombre__icontains=cliente)
    if moneda:
        transacciones = transacciones.filter(
            Q(moneda_origen__codigo=moneda) |
            Q(moneda_destino__codigo=moneda)
        )

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ganancias"

    # ==============================
    # ENCABEZADO (como reporte transacciones)
    # ==============================
    ws.merge_cells('A1:I1')
    ws['A1'] = "Reporte de Ganancias - Global Exchange"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    user = request.user
    nombre_usuario = getattr(user, "nombre", None) or getattr(user, "email", "Usuario desconocido")
    ws['A3'] = f"Generado por: {nombre_usuario}"
    ws['A4'] = f"Fecha: {now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.append([])

    # ==============================
    # ENCABEZADO TABLA
    # ==============================
    headers = ["N°", "Cliente", "Tipo Operación", "Moneda","Tasa aplicada",
               "Monto","Comisión", "Ganancia (Gs)", "Fecha"]
    ws.append(headers)

    # Estilo encabezado tabla
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    
    total_ventas = Decimal('0')
    total_compras = Decimal('0')
    total_general = Decimal('0')

    for idx, t in enumerate(transacciones, start=1):
        try:
            registro = RegistroGanancia.objects.get(transaccion=t)
            ganancia = registro.ganancia_registrada
        except RegistroGanancia.DoesNotExist:
            ganancia = Decimal('0')

        moneda_codigo = t.moneda_destino.codigo if t.tipo_operacion == "venta" else t.moneda_origen.codigo
        monto = t.monto_destino if t.tipo_operacion == "venta" else t.monto_origen
        monto_excel = Decimal(monto).quantize(Decimal("1"))
        ganancia_excel = Decimal(ganancia).quantize(Decimal("1"))

        if t.tipo_operacion == "venta":
            total_ventas += ganancia_excel
        else:
            total_compras += ganancia_excel
        total_general += ganancia_excel

        ws.append([
            idx,
            str(t.cliente),
            t.get_tipo_operacion_display(),
            moneda_codigo,
            f"{Decimal(t.tasa_cambio_aplicada):,.0f}".replace(',', '.'),
            f"{Decimal(monto_excel):,.0f}".replace(',', '.'),
            f"{Decimal(t.comision_final):,.0f}".replace(',', '.'),
            f"{Decimal(ganancia_excel):,.0f}".replace(',', '.'),
            t.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])

    # ==============================
    # TOTALES
    # ==============================
    ws.append([])
    ws.append(["", "", "", "","","", "Total Ventas", f"{total_ventas:,.0f}".replace(',', '.')])
    ws.append(["", "", "", "","","","Total Compras", f"{total_compras:,.0f}".replace(',', '.')])
    ws.append(["", "", "", "","", "","Total General", f"{total_general:,.0f}".replace(',', '.')])

    # ==============================
    # AJUSTE DE ANCHO DE COLUMNAS
    # ==============================
  # Ajustar automáticamente ancho de columnas
    for i, column_cells in enumerate(ws.columns, 1):
        column_letter = get_column_letter(i)
        if i == 1:  # Columna N°
            ws.column_dimensions[column_letter].width = 5  # ancho fijo para N°
            continue
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # ==============================
    # EXPORTAR
    # ==============================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_ganancias.xlsx"'
    wb.save(response)
    return response




from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.auth.decorators import login_required
from transacciones.models import Transaccion

# Estados definidos
ESTADO_CHOICES = [
    ('pendiente_confirmacion_pago', 'Pendiente de Confirmación de Pago'),
    ('pendiente_pago_cliente', 'Pendiente de Pago del Cliente'),
    ('pendiente_retiro_tauser', 'Pendiente de Retiro de Divisa (Tauser)'),
    ('pendiente_deposito_tauser', 'Pendiente de Depósito de Divisa (Tauser)'),
    ('procesando_acreditacion', 'Procesando Acreditación a Cliente'),
    ('pendiente_pago_stripe', 'Pendiente de Pago con Stripe'),
    ('completada', 'Completada'),
    ('cancelada', 'Cancelada'),
    ('cancelada_usuario_tasa', 'Cancelada por Usuario (Variación de Tasa)'),
    ('cancelada_tasa_expirada', 'Cancelada (Tasa Expirada)'),
    ('anulada', 'Anulada'),
    ('error', 'Error'),
]

# =========================
# PANEL PRINCIPAL DE REPORTES
# =========================
@login_required
def panel_reportes(request):
    return render(request, 'reportes/panel_reportes.html')


# =========================
# REPORTE DE TRANSACCIONES (WEB)
# =========================
@login_required
def reporte_transacciones(request):
    transacciones = Transaccion.objects.all().order_by('-fecha_creacion')

    # --- FILTROS ---
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')
    estado = request.GET.get('estado')
    cliente = request.GET.get('cliente')
    moneda = request.GET.get('moneda')

    if fecha_inicio:
        transacciones = transacciones.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        transacciones = transacciones.filter(fecha_creacion__date__lte=fecha_fin)
    if tipo:
        transacciones = transacciones.filter(tipo_operacion=tipo)
    if estado:
        transacciones = transacciones.filter(estado=estado)
    if cliente:
        transacciones = transacciones.filter(cliente__nombre__icontains=cliente) # type: ignore
    if moneda:
        transacciones = transacciones.filter(Q(moneda_origen__codigo=moneda) | Q(moneda_destino__codigo=moneda))


    monedas_disponibles = Moneda.objects.all().order_by('codigo')

    # --- PAGINACIÓN ---
    paginator = Paginator(transacciones, 10)  # 10 transacciones por página
    page_number = request.GET.get('page')     # obtener el número de página
    page_obj = paginator.get_page(page_number)  # página actual
    context = {
        'user': request.user,
        'now': now(), # type: ignore
        'transacciones': page_obj,
        'estados': ESTADO_CHOICES,
        'tipos_operacion': Transaccion.TIPO_OPERACION_CHOICES,
        'monedas': monedas_disponibles,
        'paginator': paginator,      # opcional, útil para controles de navegación
        'page_obj': page_obj,        # opcional, pero recomendable
    }
    return render(request, 'reportes/reporte_transacciones.html', context)

# =========================
# REPORTE DE TRANSACCIONES PDF
# =========================
@login_required
def reporte_transacciones_pdf(request):
    transacciones = Transaccion.objects.all().order_by('-fecha_creacion')

    # --- FILTROS ---
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')
    estado = request.GET.get('estado')
    cliente = request.GET.get('cliente')
    moneda = request.GET.get('moneda')

    if fecha_inicio:
        transacciones = transacciones.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        transacciones = transacciones.filter(fecha_creacion__date__lte=fecha_fin)
    if tipo:
        transacciones = transacciones.filter(tipo_operacion=tipo)
    if estado:
        transacciones = transacciones.filter(estado=estado)
    if cliente:
        transacciones = transacciones.filter(cliente__nombre__icontains=cliente)
    if moneda:
        transacciones = transacciones.filter(Q(moneda_origen__codigo=moneda) | Q(moneda_destino__codigo=moneda))

    # --- PDF ---
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_transacciones.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=60, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    user = request.user
    nombre_usuario = getattr(user, "nombre", None) or getattr(user, "email", "Usuario desconocido")

    elements.append(Paragraph("<b>Reporte de Transacciones - Global Exchange</b>", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generado por: {nombre_usuario}", styles["Normal"]))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elements.append(Spacer(1, 18))

    data = [["N°", "Cliente", "Tipo", "Moneda", "Monto Origen", "Monto Destino", "Estado", "Fecha"]]

    for idx, t in enumerate(transacciones, start=1):
        moneda_codigo = (
            t.moneda_origen.codigo if t.tipo_operacion.lower() == 'compra' else t.moneda_destino.codigo
        )

        data.append([
            idx,
            str(t.cliente),
            t.get_tipo_operacion_display(),
            moneda_codigo,
            f"{Decimal(t.monto_origen):,.0f}".replace(',', '.'),
            f"{Decimal(t.monto_destino):,.0f}".replace(',', '.'),
            t.get_estado_display(),
            t.fecha_creacion.strftime("%d/%m/%Y %H:%M")
        ])

    col_widths = [30, 150, 80, 50, 90, 90, 140, 90]
    table = Table(data, colWidths=col_widths, hAlign="CENTER")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("FONTSIZE", (6,1), (6,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
    ]))

    elements.append(table)
    doc.build(elements)
    return response



# =========================
# REPORTE DE TRANSACCIONES EXCEL
# =========================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

@login_required
def reporte_transacciones_excel(request):
    transacciones = Transaccion.objects.all().order_by('-fecha_creacion')

    # --- FILTROS ---
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')
    estado = request.GET.get('estado')
    cliente = request.GET.get('cliente')
    moneda = request.GET.get('moneda')

    if fecha_inicio:
        transacciones = transacciones.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        transacciones = transacciones.filter(fecha_creacion__date__lte=fecha_fin)
    if tipo:
        transacciones = transacciones.filter(tipo_operacion__iexact=tipo.upper())
    if estado:
        transacciones = transacciones.filter(estado=estado)
    if cliente:
        transacciones = transacciones.filter(cliente__nombre__icontains=cliente)
    if moneda:
        transacciones = transacciones.filter(Q(moneda_origen__codigo=moneda) | Q(moneda_destino__codigo=moneda))

    # --- Excel ---
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte_transacciones.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    ws.merge_cells('A1:H1')
    ws['A1'] = "Reporte de Transacciones - Global Exchange"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    user = request.user
    nombre_usuario = getattr(user, "nombre", None) or getattr(user, "email", "Usuario desconocido")
    ws['A3'] = f"Generado por: {nombre_usuario}"
    ws['A4'] = f"Fecha: {now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.append([])

    headers = ["#", "Cliente", "Tipo", "Moneda", "Monto Origen", "Monto Destino", "Estado", "Fecha"]
    ws.append(headers)

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, t in enumerate(transacciones, start=1):
        moneda_codigo = (
            t.moneda_origen.codigo if t.tipo_operacion.lower() == 'compra' else t.moneda_destino.codigo
        )

        ws.append([
            i,
            str(t.cliente),
            t.get_tipo_operacion_display(),
            moneda_codigo,
            f"{Decimal(t.monto_origen):,.0f}".replace(',', '.'),
            f"{Decimal(t.monto_destino):,.0f}".replace(',', '.'),
            t.get_estado_display(),
            t.fecha_creacion.strftime("%d/%m/%Y %H:%M")
        ])

    col_widths = [5, 25, 15, 10, 18, 18, 25, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    wb.save(response)
    return response
